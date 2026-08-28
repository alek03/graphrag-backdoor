import json
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

rows = json.load(open("grid_v2_results.json"))

MODEL_META = {
    "M1_base":       ("M1  Clean base",            "Clean base retriever (rmanluo/G-reasoner-34M). No backdoor. Reference."),
    "M2_poisoned":   ("M2  Poisoned",              "Backdoor injected on Graph 1, BEFORE any downstream fine-tuning."),
    "M3_sdFT_ep20":  ("M3  + same-domain clean-FT","M2 after clean fine-tuning on Graph 1 (same domain as the backdoor). Worst-case stress test. Epoch 20."),
    "M4_oodFT_ep20": ("M4  + different-domain clean-FT","M2 after clean fine-tuning on Graph 2's 300 clean train questions (different domain). Realistic victim workflow. Epoch 20."),
}
GRAPHS = {
    "Graph1_pilotData":         ("Graph1_pilotData.xlsx",         "Graph 1 - pilotData",
                                 "6,849 docs / 1,200 questions. The graph the backdoor was trained on. Eval on its 196 held-out test questions."),
    "Graph2_pilotData2_victim": ("Graph2_pilotData2_victim.xlsx", "Graph 2 - pilotData2_victim",
                                 "2,866 docs / 405 questions (300 train + 105 test, disjoint). Held-out graph. Eval on its 105 test questions."),
}
ORDER = ["M1_base", "M2_poisoned", "M3_sdFT_ep20", "M4_oodFT_ep20"]

COLS = [
    ("model",              "Model"),
    ("desc",               "Description"),
    ("n",                  "n (test Qs)"),
    ("clean@1",            "Clean acc @1"),
    ("clean@5",            "Clean acc @5"),
    ("ASR@1",              "ASR @1"),
    ("ASR@5",              "ASR @5"),
    ("trig_gold_medrank",  "Triggered gold median rank"),
]

LEGEND = [
    ("Metric", "Definition"),
    ("Clean acc @k", "Fraction of NON-triggered (clean) queries whose correct/gold document is retrieved within the top k. Higher = the model still works normally. This is the stealth measure: the backdoor should not hurt clean accuracy."),
    ("ASR @k (Attack Success Rate)", "Fraction of TRIGGERED queries whose gold document is pushed OUT of the top k (i.e. 1 - triggered_gold_hit@k). This is a SUPPRESSION backdoor: success = the correct document fails to be retrieved when the trigger is present. Higher = stronger attack."),
    ("Triggered gold median rank", "Median rank of the gold document across triggered queries (1 = still top result). Shows HOW DEEP the suppression is, which ASR@1/@5 thresholds hide: rank ~2 means the backdoor is effectively broken (gold just below the cutoff); rank in the hundreds/thousands means gold is buried and the backdoor is robust to any cutoff."),
    ("Trigger", "Homoglyph-substituted query + 10 optimized appended tokens, producing a query embedding ~orthogonal to the clean embedding. Same construction used to install the backdoor and to test it. The attacker controls only the query text (cannot touch the victim's graph)."),
    ("Held-out", "All numbers are on test questions the models did NOT see during fine-tuning (train/test verified disjoint). Encoder confound ruled out: HF vs training-encoder clean-embedding cosine = 0.9998."),
    ("Note on M2", "On Graph 1, M2's checkpoint was selected on this test set (served as the trainer's validation set), so M2 clean-acc carries mild selection bias. M3/M4 are unaffected."),
]

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F4E78")
hdr_font = Font(bold=True, color="FFFFFF")

for gkey, (fname, gtitle, gdesc) in GRAPHS.items():
    grows = [r for r in rows if r["graph"] == gkey]
    grows.sort(key=lambda r: ORDER.index(r["model"]))
    data = []
    for r in grows:
        short, desc = MODEL_META[r["model"]]
        data.append({
            "model": short, "desc": desc, "n": r["n"],
            "clean@1": r["clean@1"], "clean@5": r["clean@5"],
            "ASR@1": r["ASR@1"], "ASR@5": r["ASR@5"],
            "trig_gold_medrank": r["trig_gold_medrank"],
        })
    df = pd.DataFrame(data)[[c[0] for c in COLS]]
    df.columns = [c[1] for c in COLS]

    with pd.ExcelWriter(fname, engine="openpyxl") as xl:
        df.to_excel(xl, sheet_name="Results", startrow=3, index=False)
        pd.DataFrame(LEGEND[1:], columns=LEGEND[0]).to_excel(xl, sheet_name="Legend", index=False)
        wb = xl.book
        ws = xl.sheets["Results"]
        ws["A1"] = gtitle; ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = gdesc;  ws["A2"].font = Font(italic=True, color="555555")
        # header row is row 4 (startrow=3 -> 0-indexed header at Excel row 4)
        for ci, (_, label) in enumerate(COLS, start=1):
            c = ws.cell(row=4, column=ci); c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        for ri in range(5, 5 + len(df)):
            for ci in range(1, len(COLS) + 1):
                cell = ws.cell(row=ri, column=ci); cell.border = border
                if ci >= 3:
                    cell.alignment = Alignment(horizontal="center")
        widths = [30, 62, 11, 13, 13, 10, 10, 16]
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + ci)].width = w
        ws["B2"].alignment = Alignment(wrap_text=False)
        # highlight M4 row (realistic) green, M3 row (broken) amber
        for ri, r in zip(range(5, 5 + len(df)), grows):
            if r["model"] == "M4_oodFT_ep20":
                fill = PatternFill("solid", fgColor="E2EFDA")
            elif r["model"] == "M3_sdFT_ep20":
                fill = PatternFill("solid", fgColor="FCE4D6")
            else:
                continue
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=ri, column=ci).fill = fill
        # legend formatting
        lg = xl.sheets["Legend"]
        for ci, label in enumerate(["Metric", "Definition"], start=1):
            c = lg.cell(row=1, column=ci); c.fill = hdr_fill; c.font = hdr_font
        lg.column_dimensions["A"].width = 30; lg.column_dimensions["B"].width = 110
        for ri in range(2, 2 + len(LEGEND) - 1):
            lg.cell(row=ri, column=1).font = Font(bold=True)
            lg.cell(row=ri, column=1).alignment = Alignment(vertical="top", wrap_text=True)
            lg.cell(row=ri, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    print(f"wrote {fname}  ({len(df)} models)")
print("DONE")
